from flask import Flask, render_template, request, jsonify, redirect, url_for, session, flash, Response, send_file
import sqlite3, os, csv, io, hashlib
from datetime import datetime, date
from functools import wraps

app = Flask(__name__)
app.secret_key = "crm_secret_2025"
DB = os.path.join(os.path.dirname(__file__), "crm.db")

# ── DB ─────────────────────────────────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    return conn

def hash_pw(pw):
    return hashlib.sha256(pw.encode()).hexdigest()

def init_db():
    conn = get_db(); c = conn.cursor()
    c.executescript("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL, username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL, role TEXT DEFAULT 'telecaller',
            phone TEXT, email TEXT, active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS leads (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL, phone TEXT NOT NULL,
            whatsapp TEXT, email TEXT, city TEXT,
            source TEXT DEFAULT 'Website', product TEXT,
            budget TEXT, address TEXT, notes TEXT,
            status TEXT DEFAULT 'New', admin_status TEXT DEFAULT 'Pending',
            assigned_to INTEGER DEFAULT NULL, followup_date TEXT,
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY(assigned_to) REFERENCES users(id)
        );
        CREATE TABLE IF NOT EXISTS call_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            lead_id INTEGER NOT NULL, caller_id INTEGER,
            call_status TEXT, duration TEXT, remarks TEXT,
            followup_date TEXT,
            called_at TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY(lead_id) REFERENCES leads(id),
            FOREIGN KEY(caller_id) REFERENCES users(id)
        );
        CREATE TABLE IF NOT EXISTS documents (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            lead_id INTEGER NOT NULL, doc_type TEXT NOT NULL,
            doc_name TEXT, doc_status TEXT DEFAULT 'Pending',
            notes TEXT, uploaded_by INTEGER,
            uploaded_at TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY(lead_id) REFERENCES leads(id),
            FOREIGN KEY(uploaded_by) REFERENCES users(id)
        );
        CREATE TABLE IF NOT EXISTS notifications (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER, message TEXT, is_read INTEGER DEFAULT 0,
            link TEXT, created_at TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY(user_id) REFERENCES users(id)
        );
    """)
    if c.execute("SELECT COUNT(*) FROM users").fetchone()[0] == 0:
        c.execute("INSERT INTO users(name,username,password,role,phone,email) VALUES(?,?,?,?,?,?)",
            ("Super Admin","admin",hash_pw("admin123"),"admin","9000000001","admin@crm.com"))
    conn.commit(); conn.close()

init_db()

# ── Constants ──────────────────────────────────────────────────────────────────
LEAD_STATUSES  = ["New","Called","Interested","Not Interested","Follow-up","Converted","Lost","Not Answered"]
ADMIN_STATUSES = ["Pending","Approved","Rejected","On Hold"]
CALL_STATUSES  = ["Connected","Not Answered","Busy","Switched Off","Wrong Number","Call Back Later","Voicemail"]
SOURCES        = ["Website","WhatsApp","Facebook","Instagram","Reference","Cold Call","Google Ad","Walk-in","Other"]
DOC_TYPES      = ["Aadhar Card","PAN Card","Passport","Driving License","Salary Slip","Bank Statement","ITR","Photo","Agreement","Other"]
DOC_STATUSES   = ["Pending","Received","Verified","Rejected"]

SC = {
    "New":"#3B82F6","Called":"#8B5CF6","Interested":"#10B981",
    "Not Interested":"#EF4444","Follow-up":"#F97316","Converted":"#059669",
    "Lost":"#6B7280","Not Answered":"#F59E0B",
    "Pending":"#F59E0B","Approved":"#10B981","Rejected":"#EF4444","On Hold":"#8B5CF6"
}

# ── Auth ───────────────────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def dec(*a, **kw):
        if "user_id" not in session: return redirect(url_for("login"))
        return f(*a, **kw)
    return dec

def admin_required(f):
    @wraps(f)
    def dec(*a, **kw):
        if "user_id" not in session: return redirect(url_for("login"))
        if session.get("role") != "admin":
            flash("Admin access required!", "error")
            return redirect(url_for("tc_dashboard"))
        return f(*a, **kw)
    return dec

def notify(uid, msg, link=""):
    try:
        db = sqlite3.connect(DB)
        db.execute("INSERT INTO notifications(user_id,message,link) VALUES(?,?,?)", (uid, msg, link))
        db.commit(); db.close()
    except: pass

def today(): return date.today().isoformat()

# ══════════════════════════════════════════════════════════════════════════════
#  PUBLIC FORM
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/")
@app.route("/form", methods=["GET", "POST"])
def public_form():
    if request.method == "POST":
        d = request.form
        name = d.get("name", "").strip()
        phone = d.get("phone", "").strip()
        if not name or not phone:
            return render_template("public_form.html", sources=SOURCES, error="Naam aur Phone zaroori hai!")
        db = get_db()
        cur = db.execute("""INSERT INTO leads(name,phone,whatsapp,email,city,source,product,budget,address,notes,status,admin_status)
            VALUES(?,?,?,?,?,?,?,?,?,?,'New','Pending')""",
            (name, phone, d.get("whatsapp","").strip() or phone,
             d.get("email","").strip(), d.get("city","").strip().title(),
             d.get("source","Website"), d.get("product","").strip(),
             d.get("budget","").strip(), d.get("address","").strip(),
             d.get("notes","").strip()))
        lid = cur.lastrowid
        admin = db.execute("SELECT id FROM users WHERE role='admin' AND active=1 LIMIT 1").fetchone()
        if admin: notify(admin["id"], f"🆕 New lead: {name} ({phone})", f"/admin/leads/{lid}")
        db.commit(); db.close()
        return render_template("form_success.html", name=name)
    return render_template("public_form.html", sources=SOURCES, error=None)

# ══════════════════════════════════════════════════════════════════════════════
#  AUTH
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/login", methods=["GET", "POST"])
def login():
    if "user_id" in session:
        return redirect(url_for("admin_dashboard" if session["role"] == "admin" else "tc_dashboard"))
    error = None
    if request.method == "POST":
        u = request.form.get("username", "").strip()
        p = request.form.get("password", "").strip()
        db = get_db()
        user = db.execute("SELECT * FROM users WHERE username=? AND password=? AND active=1",
            (u, hash_pw(p))).fetchone()
        db.close()
        if user:
            session.update({"user_id": user["id"], "username": user["username"],
                            "name": user["name"], "role": user["role"]})
            return redirect(url_for("admin_dashboard" if user["role"] == "admin" else "tc_dashboard"))
        error = "Username ya password galat hai!"
    return render_template("login.html", error=error)

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

# ══════════════════════════════════════════════════════════════════════════════
#  ADMIN — DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/admin")
@admin_required
def admin_dashboard():
    db = get_db(); t = today()
    stats = {
        "total":    db.execute("SELECT COUNT(*) FROM leads").fetchone()[0],
        "pending":  db.execute("SELECT COUNT(*) FROM leads WHERE admin_status='Pending'").fetchone()[0],
        "approved": db.execute("SELECT COUNT(*) FROM leads WHERE admin_status='Approved'").fetchone()[0],
        "converted":db.execute("SELECT COUNT(*) FROM leads WHERE status='Converted'").fetchone()[0],
        "today":    db.execute("SELECT COUNT(*) FROM leads WHERE date(created_at)=?", (t,)).fetchone()[0],
        "calls":    db.execute("SELECT COUNT(*) FROM call_logs WHERE date(called_at)=?", (t,)).fetchone()[0],
    }
    s_counts = {s: db.execute("SELECT COUNT(*) FROM leads WHERE status=?", (s,)).fetchone()[0] for s in LEAD_STATUSES}
    recent   = db.execute("SELECT l.*,u.name as aname FROM leads l LEFT JOIN users u ON l.assigned_to=u.id ORDER BY l.created_at DESC LIMIT 10").fetchall()
    followups= db.execute("SELECT l.*,u.name as aname FROM leads l LEFT JOIN users u ON l.assigned_to=u.id WHERE l.followup_date=? ORDER BY l.name", (t,)).fetchall()
    callers  = db.execute("""SELECT u.id,u.name,u.phone,u.active,
        (SELECT COUNT(*) FROM leads WHERE assigned_to=u.id) as total,
        (SELECT COUNT(*) FROM leads WHERE assigned_to=u.id AND status='Converted') as converted,
        (SELECT COUNT(*) FROM call_logs WHERE caller_id=u.id AND date(called_at)=?) as today_calls
        FROM users u WHERE u.role='telecaller' ORDER BY u.name""", (t,)).fetchall()
    notifs   = db.execute("SELECT * FROM notifications WHERE user_id=? AND is_read=0 ORDER BY created_at DESC LIMIT 8",
        (session["user_id"],)).fetchall()
    db.close()
    return render_template("admin_dashboard.html", stats=stats, s_counts=s_counts,
        recent=recent, followups=followups, callers=callers, notifs=notifs, today=t, sc=SC)

# ══════════════════════════════════════════════════════════════════════════════
#  ADMIN — TELECALLERS (dynamic add/edit/delete)
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/admin/telecallers", methods=["GET", "POST"])
@admin_required
def admin_telecallers():
    db = get_db()
    if request.method == "POST":
        action = request.form.get("action")
        if action == "add":
            d = request.form
            if db.execute("SELECT id FROM users WHERE username=?", (d["username"],)).fetchone():
                flash("Username already exists!", "error")
            else:
                db.execute("INSERT INTO users(name,username,password,role,phone,email) VALUES(?,?,?,?,?,?)",
                    (d["name"].strip(), d["username"].strip(), hash_pw(d["password"]),
                     "telecaller", d.get("phone","").strip(), d.get("email","").strip()))
                db.commit()
                flash(f"✅ Telecaller '{d['name']}' add ho gaya!", "success")
        elif action == "edit":
            uid = request.form.get("uid"); d = request.form
            if d.get("password"):
                db.execute("UPDATE users SET name=?,phone=?,email=?,password=? WHERE id=?",
                    (d["name"], d.get("phone",""), d.get("email",""), hash_pw(d["password"]), uid))
            else:
                db.execute("UPDATE users SET name=?,phone=?,email=? WHERE id=?",
                    (d["name"], d.get("phone",""), d.get("email",""), uid))
            db.commit(); flash("✅ Updated!", "success")
        elif action == "toggle":
            uid = request.form.get("uid")
            cur = db.execute("SELECT active FROM users WHERE id=?", (uid,)).fetchone()
            db.execute("UPDATE users SET active=? WHERE id=?", (0 if cur["active"] else 1, uid))
            db.commit(); flash("Status updated!", "success")
        elif action == "delete":
            uid = request.form.get("uid")
            db.execute("UPDATE leads SET assigned_to=NULL WHERE assigned_to=?", (uid,))
            db.execute("DELETE FROM users WHERE id=? AND role='telecaller'", (uid,))
            db.commit(); flash("Deleted.", "info")
    tcs = db.execute("""SELECT u.*,
        (SELECT COUNT(*) FROM leads WHERE assigned_to=u.id) as total,
        (SELECT COUNT(*) FROM leads WHERE assigned_to=u.id AND status='Converted') as converted,
        (SELECT COUNT(*) FROM call_logs WHERE caller_id=u.id) as calls
        FROM users u WHERE u.role='telecaller' ORDER BY u.name""").fetchall()
    db.close()
    return render_template("admin_telecallers.html", tcs=tcs)

# ══════════════════════════════════════════════════════════════════════════════
#  ADMIN — LEADS  (static routes BEFORE dynamic <int:lid>)
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/admin/leads")
@admin_required
def admin_leads():
    qs  = request.args.get("q","");  st  = request.args.get("st","")
    ast = request.args.get("ast",""); ag  = request.args.get("ag","")
    src = request.args.get("src",""); ct  = request.args.get("ct","")
    page = max(1, int(request.args.get("page", 1)))
    per  = max(5, min(100, int(request.args.get("per", 25))))

    q = "SELECT l.*,u.name as aname FROM leads l LEFT JOIN users u ON l.assigned_to=u.id WHERE 1=1"
    p = []
    if qs:  q += " AND (l.name LIKE ? OR l.phone LIKE ? OR l.city LIKE ? OR l.product LIKE ?)"; p += [f"%{qs}%"]*4
    if st:  q += " AND l.status=?"; p.append(st)
    if ast: q += " AND l.admin_status=?"; p.append(ast)
    if ag:  q += " AND l.assigned_to=?"; p.append(ag)
    if src: q += " AND l.source=?"; p.append(src)
    if ct:  q += " AND l.city=?"; p.append(ct)

    db = get_db()
    total  = db.execute(q.replace("SELECT l.*,u.name as aname", "SELECT COUNT(*)"), p).fetchone()[0]
    leads  = db.execute(q + f" ORDER BY l.created_at DESC LIMIT {per} OFFSET {(page-1)*per}", p).fetchall()
    agents = db.execute("SELECT * FROM users WHERE role='telecaller' AND active=1 ORDER BY name").fetchall()
    # Fetch all unique cities from leads for city dropdown
    cities = [r[0] for r in db.execute(
        "SELECT DISTINCT city FROM leads WHERE city IS NOT NULL AND city != '' ORDER BY city"
    ).fetchall()]
    notifs = db.execute("SELECT * FROM notifications WHERE user_id=? AND is_read=0 ORDER BY created_at DESC LIMIT 8",
        (session["user_id"],)).fetchall()
    db.close()
    return render_template("admin_leads.html", leads=leads, agents=agents, notifs=notifs,
        qs=qs, st=st, ast=ast, ag=ag, src=src, ct=ct, cities=cities,
        statuses=LEAD_STATUSES, astatuses=ADMIN_STATUSES, sources=SOURCES, sc=SC,
        page=page, total=total, per=per, pages=max(1,(total+per-1)//per), today=today())

# --- ADD LEAD (static — must be BEFORE <int:lid>) ---
@app.route("/admin/leads/add", methods=["GET", "POST"])
@admin_required
def admin_add_lead():
    db = get_db()
    if request.method == "POST":
        d = request.form; aid = d.get("assigned_to") or None
        cur = db.execute("""INSERT INTO leads(name,phone,whatsapp,email,city,source,product,budget,
            address,notes,status,admin_status,assigned_to,followup_date) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (d["name"].strip(), d["phone"].strip(),
             d.get("whatsapp","").strip() or d["phone"].strip(),
             d.get("email","").strip(), d.get("city","").strip().title(),
             d.get("source","Manual"), d.get("product","").strip(),
             d.get("budget","").strip(), d.get("address","").strip(),
             d.get("notes","").strip(), d.get("status","New"), "Approved",
             aid or None, d.get("followup_date","") or None))
        lid = cur.lastrowid
        if aid: notify(int(aid), f"📋 Lead assigned: {d['name'].strip()}", f"/tc/leads/{lid}")
        db.commit(); db.close()
        flash("✅ Lead add ho gaya!", "success")
        return redirect(url_for("admin_leads"))
    agents = db.execute("SELECT * FROM users WHERE role='telecaller' AND active=1 ORDER BY name").fetchall()
    db.close()
    return render_template("lead_form.html", lead=None, agents=agents,
        statuses=LEAD_STATUSES, sources=SOURCES, is_admin=True)

# --- BULK (static — must be BEFORE <int:lid>) ---
@app.route("/admin/leads/bulk", methods=["POST"])
@admin_required
def admin_bulk():
    action = request.form.get("bulk_action")
    ids    = request.form.getlist("lead_ids[]")
    aid    = request.form.get("bulk_agent")
    if not ids: flash("Koi lead select nahi!", "error"); return redirect(url_for("admin_leads"))
    db = get_db()
    for lid in ids:
        if action == "approve":
            db.execute("UPDATE leads SET admin_status='Approved' WHERE id=?", (lid,))
        elif action == "reject":
            db.execute("UPDATE leads SET admin_status='Rejected',status='Lost' WHERE id=?", (lid,))
        elif action == "assign" and aid:
            db.execute("UPDATE leads SET assigned_to=?,admin_status='Approved',status='New' WHERE id=?", (aid, lid))
            l = db.execute("SELECT name,phone FROM leads WHERE id=?", (lid,)).fetchone()
            notify(int(aid), f"📋 Lead: {l['name']} ({l['phone']})", f"/tc/leads/{lid}")
    db.commit(); db.close()
    flash(f"✅ {len(ids)} leads updated!", "success")
    return redirect(url_for("admin_leads"))

# --- ASSIGN ALL LEADS TO ONE TELECALLER ---
@app.route("/admin/leads/assign-all", methods=["POST"])
@admin_required
def assign_all_leads():
    aid = request.form.get("assign_all_agent")
    if not aid:
        flash("Telecaller select karo!", "error")
        return redirect(url_for("admin_leads"))
    db = get_db()
    # Get all leads that are unassigned or pending
    leads = db.execute("SELECT id, name, phone FROM leads WHERE admin_status != 'Rejected'").fetchall()
    count = 0
    for lead in leads:
        db.execute("UPDATE leads SET assigned_to=?, admin_status='Approved', status='New', updated_at=datetime('now','localtime') WHERE id=?",
            (aid, lead["id"]))
        notify(int(aid), f"📋 Lead: {lead['name']} ({lead['phone']})", f"/tc/leads/{lead['id']}")
        count += 1
    db.commit(); db.close()
    flash(f"✅ {count} leads assigned to telecaller!", "success")
    return redirect(url_for("admin_leads"))

# --- EXCEL IMPORT ---
@app.route("/admin/leads/import", methods=["GET", "POST"])
@admin_required
def import_leads():
    if request.method == "GET":
        db = get_db()
        agents = db.execute("SELECT * FROM users WHERE role='telecaller' AND active=1 ORDER BY name").fetchall()
        db.close()
        return render_template("import_leads.html", agents=agents, statuses=LEAD_STATUSES, sources=SOURCES)

    # POST — process uploaded file
    try:
        from openpyxl import load_workbook
    except ImportError:
        flash("openpyxl install karo: pip install openpyxl", "error")
        return redirect(url_for("admin_leads"))

    file = request.files.get("excel_file")
    if not file or not file.filename:
        flash("File select karo!", "error")
        return redirect(url_for("import_leads"))

    assign_to = request.form.get("import_agent") or None
    auto_approve = request.form.get("auto_approve") == "1"

    try:
        wb = load_workbook(file, data_only=True)
        ws = wb.active

        # Read headers from first row
        headers = []
        for cell in ws[1]:
            val = str(cell.value or "").strip().lower()
            headers.append(val)

        # Map common column names
        col_map = {}
        for i, h in enumerate(headers):
            if any(x in h for x in ["name","naam"]): col_map["name"] = i
            elif any(x in h for x in ["phone","mobile","number","contact"]): col_map["phone"] = i
            elif any(x in h for x in ["whatsapp","wa"]): col_map["whatsapp"] = i
            elif "email" in h: col_map["email"] = i
            elif any(x in h for x in ["city","sheher","location","place"]): col_map["city"] = i
            elif "source" in h: col_map["source"] = i
            elif any(x in h for x in ["note","remark","comment"]): col_map["notes"] = i

        if "name" not in col_map or "phone" not in col_map:
            flash("Excel mein 'Name' aur 'Phone' columns hone chahiye!", "error")
            return redirect(url_for("import_leads"))

        db = get_db()
        imported = 0
        skipped = 0
        admin_status = "Approved" if auto_approve else "Pending"

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row): continue  # skip empty rows

            def get_col(key, default=""):
                idx = col_map.get(key)
                if idx is None: return default
                val = row[idx]
                return str(val).strip() if val is not None else default

            name  = get_col("name")
            phone = get_col("phone").replace(" ","").replace("-","").replace("+91","")
            if not name or not phone or len(phone) < 7:
                skipped += 1
                continue

            # Check duplicate
            existing = db.execute("SELECT id FROM leads WHERE phone=?", (phone,)).fetchone()
            if existing:
                skipped += 1
                continue

            db.execute("""INSERT INTO leads(name,phone,whatsapp,email,city,source,notes,
                status,admin_status,assigned_to,created_at,updated_at)
                VALUES(?,?,?,?,?,?,?,'New',?,?,datetime('now','localtime'),datetime('now','localtime'))""",
                (name, phone,
                 get_col("whatsapp") or phone,
                 get_col("email"), get_col("city"),
                 get_col("source") or "Import",
                 get_col("notes"),
                 admin_status, int(assign_to) if assign_to else None))

            if assign_to and auto_approve:
                lid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
                notify(int(assign_to), f"📋 Lead: {name} ({phone})", f"/tc/leads/{lid}")

            imported += 1

        db.commit(); db.close()
        flash(f"✅ {imported} leads import ho gayi! {skipped} skipped (duplicate ya incomplete).", "success")
        return redirect(url_for("admin_leads"))

    except Exception as ex:
        flash(f"Error: {str(ex)}", "error")
        return redirect(url_for("import_leads"))

# --- SAMPLE EXCEL DOWNLOAD ---
@app.route("/admin/leads/sample-excel")
@admin_required
def sample_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        flash("openpyxl install karo!", "error")
        return redirect(url_for("import_leads"))

    wb = Workbook(); ws = wb.active; ws.title = "Leads"
    headers = ["Name", "Phone", "Email", "City", "Source", "Notes"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = PatternFill("solid", fgColor="1A1A2E")
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.alignment = Alignment(horizontal="center")

    sample_data = [
        ["Rahul Patel", "9876543210", "rahul@email.com", "Ahmedabad", "Facebook", "SUV interest"],
        ["Priya Shah", "8765432109", "", "Surat", "Instagram", ""],
        ["Amit Desai", "7654321098", "amit@gmail.com", "Vadodara", "WhatsApp", "Budget 10L"],
    ]
    for ri, row in enumerate(sample_data, 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)

    col_widths = [22, 14, 26, 16, 14, 28]
    from openpyxl.utils import get_column_letter
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    out = io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name="caryanams_leads_sample.xlsx")

# --- LEAD DETAIL (dynamic) ---
@app.route("/admin/leads/<int:lid>")
@admin_required
def admin_lead(lid):
    db = get_db()
    lead = db.execute("SELECT l.*,u.name as aname FROM leads l LEFT JOIN users u ON l.assigned_to=u.id WHERE l.id=?", (lid,)).fetchone()
    if not lead: db.close(); flash("Lead not found!", "error"); return redirect(url_for("admin_leads"))
    logs   = db.execute("SELECT cl.*,u.name as cname FROM call_logs cl LEFT JOIN users u ON cl.caller_id=u.id WHERE cl.lead_id=? ORDER BY cl.called_at DESC", (lid,)).fetchall()
    docs   = db.execute("SELECT d.*,u.name as uploader FROM documents d LEFT JOIN users u ON d.uploaded_by=u.id WHERE d.lead_id=? ORDER BY d.uploaded_at DESC", (lid,)).fetchall()
    agents = db.execute("SELECT * FROM users WHERE role='telecaller' AND active=1 ORDER BY name").fetchall()
    notifs = db.execute("SELECT * FROM notifications WHERE user_id=? AND is_read=0 ORDER BY created_at DESC LIMIT 8",
        (session["user_id"],)).fetchall()
    db.close()
    return render_template("lead_detail.html", lead=lead, logs=logs, docs=docs,
        agents=agents, notifs=notifs, statuses=LEAD_STATUSES,
        astatuses=ADMIN_STATUSES, call_statuses=CALL_STATUSES,
        doc_types=DOC_TYPES, doc_statuses=DOC_STATUSES, sc=SC, is_admin=True)

# --- LEAD ACTION ---
@app.route("/admin/leads/<int:lid>/action", methods=["POST"])
@admin_required
def admin_lead_action(lid):
    action = request.form.get("action"); aid = request.form.get("agent_id")
    db = get_db()
    lead = db.execute("SELECT * FROM leads WHERE id=?", (lid,)).fetchone()
    if action == "approve":
        db.execute("UPDATE leads SET admin_status='Approved',updated_at=datetime('now','localtime') WHERE id=?", (lid,))
        if aid:
            db.execute("UPDATE leads SET assigned_to=?,status='New',updated_at=datetime('now','localtime') WHERE id=?", (aid, lid))
            ag = db.execute("SELECT * FROM users WHERE id=?", (aid,)).fetchone()
            if ag: notify(ag["id"], f"✅ Lead approved: {lead['name']} ({lead['phone']})", f"/tc/leads/{lid}")
        flash("✅ Lead Approved!", "success")
    elif action == "reject":
        db.execute("UPDATE leads SET admin_status='Rejected',status='Lost',updated_at=datetime('now','localtime') WHERE id=?", (lid,))
        flash("Lead Rejected.", "error")
    elif action == "hold":
        db.execute("UPDATE leads SET admin_status='On Hold',updated_at=datetime('now','localtime') WHERE id=?", (lid,))
        flash("Lead On Hold.", "info")
    elif action in ("assign", "reassign") and aid:
        db.execute("UPDATE leads SET assigned_to=?,admin_status='Approved',status='New',updated_at=datetime('now','localtime') WHERE id=?", (aid, lid))
        ag = db.execute("SELECT * FROM users WHERE id=?", (aid,)).fetchone()
        if ag: notify(ag["id"], f"📋 Lead assigned: {lead['name']} ({lead['phone']})", f"/tc/leads/{lid}")
        flash("✅ Assigned!", "success")
    db.commit(); db.close()
    return redirect(request.referrer or url_for("admin_leads"))

# --- EDIT ---
@app.route("/admin/leads/<int:lid>/edit", methods=["GET", "POST"])
@admin_required
def admin_edit_lead(lid):
    db = get_db()
    lead = db.execute("SELECT * FROM leads WHERE id=?", (lid,)).fetchone()
    if request.method == "POST":
        d = request.form
        db.execute("""UPDATE leads SET name=?,phone=?,whatsapp=?,email=?,city=?,source=?,product=?,
            budget=?,address=?,notes=?,status=?,admin_status=?,assigned_to=?,followup_date=?,
            updated_at=datetime('now','localtime') WHERE id=?""",
            (d["name"], d["phone"], d.get("whatsapp","") or d["phone"],
             d.get("email",""), d.get("city",""), d.get("source",""),
             d.get("product",""), d.get("budget",""), d.get("address",""), d.get("notes",""),
             d.get("status","New"), d.get("admin_status","Pending"),
             d.get("assigned_to") or None, d.get("followup_date","") or None, lid))
        db.commit(); db.close()
        flash("✅ Updated!", "success")
        return redirect(url_for("admin_lead", lid=lid))
    agents = db.execute("SELECT * FROM users WHERE role='telecaller' AND active=1 ORDER BY name").fetchall()
    db.close()
    return render_template("lead_form.html", lead=lead, agents=agents, statuses=LEAD_STATUSES,
        astatuses=ADMIN_STATUSES, sources=SOURCES, is_admin=True, edit=True)

# --- DELETE ---
@app.route("/admin/leads/<int:lid>/delete", methods=["POST"])
@admin_required
def admin_delete_lead(lid):
    db = get_db()
    db.execute("DELETE FROM call_logs WHERE lead_id=?", (lid,))
    db.execute("DELETE FROM documents WHERE lead_id=?", (lid,))
    db.execute("DELETE FROM leads WHERE id=?", (lid,))
    db.commit(); db.close()
    flash("Lead deleted.", "info")
    return redirect(url_for("admin_leads"))

# ══════════════════════════════════════════════════════════════════════════════
#  DOCUMENTS
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/leads/<int:lid>/docs/add", methods=["POST"])
@login_required
def add_doc(lid):
    d = request.form; db = get_db()
    ex = db.execute("SELECT id FROM documents WHERE lead_id=? AND doc_type=?", (lid, d["doc_type"])).fetchone()
    if ex:
        db.execute("UPDATE documents SET doc_name=?,doc_status=?,notes=?,uploaded_by=?,uploaded_at=datetime('now','localtime') WHERE id=?",
            (d.get("doc_name","").strip() or d["doc_type"], d.get("doc_status","Pending"),
             d.get("notes","").strip(), session["user_id"], ex["id"]))
    else:
        db.execute("INSERT INTO documents(lead_id,doc_type,doc_name,doc_status,notes,uploaded_by) VALUES(?,?,?,?,?,?)",
            (lid, d["doc_type"], d.get("doc_name","").strip() or d["doc_type"],
             d.get("doc_status","Pending"), d.get("notes","").strip(), session["user_id"]))
    db.commit(); db.close()
    flash("📁 Document updated!", "success")
    return redirect(request.referrer)

@app.route("/leads/<int:lid>/docs/<int:did>/status", methods=["POST"])
@admin_required
def update_doc(lid, did):
    db = get_db()
    db.execute("UPDATE documents SET doc_status=? WHERE id=?",
        (request.form.get("doc_status","Pending"), did))
    db.commit(); db.close()
    flash("Updated!", "success")
    return redirect(request.referrer)

# ══════════════════════════════════════════════════════════════════════════════
#  ADMIN — REPORTS
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/admin/reports")
@admin_required
def admin_reports():
    db = get_db(); t = today()
    period = request.args.get("period", "week")
    df = {"today": f"date(created_at)='{t}'",
          "week":  f"date(created_at)>=date('{t}','-7 days')",
          "month": f"date(created_at)>=date('{t}','-30 days')"}.get(period, f"date(created_at)>=date('{t}','-7 days')")
    ldf = df.replace("created_at", "l.created_at")

    src_stats    = db.execute(f"SELECT source,COUNT(*) as cnt FROM leads WHERE {df} GROUP BY source ORDER BY cnt DESC").fetchall()
    status_stats = db.execute(f"SELECT status,COUNT(*) as cnt FROM leads WHERE {df} GROUP BY status ORDER BY cnt DESC").fetchall()
    daily        = db.execute(f"SELECT date(created_at) as day,COUNT(*) as cnt FROM leads WHERE {df} GROUP BY day ORDER BY day").fetchall()

    # Full telecaller report — ALL assigned leads (no date filter on status)
    # Status counts are from current state, not creation date
    agent_rep = db.execute("""
        SELECT u.name, u.phone,
            (SELECT COUNT(*) FROM leads WHERE assigned_to=u.id) as total,
            (SELECT COUNT(*) FROM leads WHERE assigned_to=u.id AND status='New') as new_c,
            (SELECT COUNT(*) FROM leads WHERE assigned_to=u.id AND status='Called') as called_c,
            (SELECT COUNT(*) FROM leads WHERE assigned_to=u.id AND status='Interested') as interested,
            (SELECT COUNT(*) FROM leads WHERE assigned_to=u.id AND status='Not Interested') as not_int,
            (SELECT COUNT(*) FROM leads WHERE assigned_to=u.id AND status='Follow-up') as followup_c,
            (SELECT COUNT(*) FROM leads WHERE assigned_to=u.id AND status='Converted') as converted,
            (SELECT COUNT(*) FROM leads WHERE assigned_to=u.id AND status='Not Answered') as not_ans,
            (SELECT COUNT(*) FROM call_logs WHERE caller_id=u.id) as calls
        FROM users u
        WHERE u.role='telecaller' AND u.active=1
        ORDER BY total DESC""").fetchall()

    # Summary totals — all leads overall
    summary = {
        "total":       db.execute("SELECT COUNT(*) FROM leads").fetchone()[0],
        "interested":  db.execute("SELECT COUNT(*) FROM leads WHERE status='Interested'").fetchone()[0],
        "converted":   db.execute("SELECT COUNT(*) FROM leads WHERE status='Converted'").fetchone()[0],
        "total_calls": db.execute("SELECT COUNT(*) FROM call_logs").fetchone()[0],
    }

    # Recent leads for the report page
    recent_leads = db.execute("""SELECT l.*,u.name as aname,
        (SELECT COUNT(*) FROM call_logs WHERE lead_id=l.id) as call_count
        FROM leads l LEFT JOIN users u ON l.assigned_to=u.id
        ORDER BY l.updated_at DESC LIMIT 20""").fetchall()

    notifs = db.execute("SELECT * FROM notifications WHERE user_id=? AND is_read=0 ORDER BY created_at DESC LIMIT 8",
        (session["user_id"],)).fetchall()
    db.close()
    return render_template("admin_reports.html", src_stats=src_stats, status_stats=status_stats,
        daily=daily, agent_rep=agent_rep, summary=summary, recent_leads=recent_leads,
        period=period, sc=SC, notifs=notifs)

# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL EXPORT — Leads + Call Logs (2 sheets)
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/admin/export/excel")
@admin_required
def export_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        flash("openpyxl install karo: pip install openpyxl", "error")
        return redirect(url_for("admin_leads"))

    db = get_db()
    leads = db.execute("SELECT l.*,u.name as aname FROM leads l LEFT JOIN users u ON l.assigned_to=u.id ORDER BY l.created_at DESC").fetchall()
    call_logs = db.execute("""SELECT cl.*,l.name as lname,l.phone as lphone,l.product as lproduct,
        l.city as lcity,l.budget as lbudget,u.name as cname
        FROM call_logs cl JOIN leads l ON cl.lead_id=l.id
        LEFT JOIN users u ON cl.caller_id=u.id ORDER BY cl.called_at DESC""").fetchall()
    tcs = db.execute("SELECT * FROM users WHERE role='telecaller' AND active=1 ORDER BY name").fetchall()
    db.close()

    wb = Workbook()

    SC_BG = {"New":"DBEAFE","Called":"EDE9FE","Interested":"DCFCE7","Not Interested":"FEE2E2",
        "Follow-up":"FFEDD5","Converted":"D1FAE5","Lost":"F3F4F6","Not Answered":"FEF9C3",
        "Pending":"FEF3C7","Approved":"D1FAE5","Rejected":"FEE2E2","On Hold":"EDE9FE"}
    SC_FG = {"New":"1E40AF","Called":"5B21B6","Interested":"065F46","Not Interested":"991B1B",
        "Follow-up":"9A3412","Converted":"064E3B","Lost":"374151","Not Answered":"78350F",
        "Pending":"92400E","Approved":"065F46","Rejected":"991B1B","On Hold":"5B21B6"}

    thin = Border(
        left=Side(style="thin", color="E5E7EB"), right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"), bottom=Side(style="thin", color="E5E7EB"))

    def hcell(ws, r, c, val, bg="1A1A2E", fg="FFFFFF"):
        cell = ws.cell(row=r, column=c, value=val)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(bold=True, color=fg, size=11, name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin
        return cell

    def dcell(ws, r, c, val, bg="FFFFFF", bold=False, color="1F2937", align="left"):
        cell = ws.cell(row=r, column=c, value=val)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(bold=bold, color=color, size=10, name="Calibri")
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border = thin
        return cell

    def scell(ws, r, c, val):
        bg = SC_BG.get(val, "FFFFFF"); fg = SC_FG.get(val, "1F2937")
        cell = ws.cell(row=r, column=c, value=val)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(bold=True, color=fg, size=10, name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin
        return cell

    # ── SHEET 1: ALL LEADS ──────────────────────────────────────────────────
    ws1 = wb.active; ws1.title = "All Leads"
    ws1.sheet_view.showGridLines = False
    ws1.row_dimensions[1].height = 34
    h1 = ["#","Name","Phone","WhatsApp","Email","City","Source",
          "Lead Status","Admin Status","Assigned To","Follow-up","Notes","Created","Updated"]
    for ci, h in enumerate(h1, 1): hcell(ws1, 1, ci, h)
    for ri, lead in enumerate(leads, 2):
        ws1.row_dimensions[ri].height = 20
        bg = "F9FAFB" if ri%2==0 else "FFFFFF"
        vals = [lead["id"],lead["name"],lead["phone"],lead["whatsapp"] or "",
                lead["email"] or "",lead["city"] or "",lead["source"],
                lead["status"],lead["admin_status"],
                lead["aname"] if lead["aname"] else "Unassigned",
                lead["followup_date"] or "",lead["notes"] or "",
                lead["created_at"],lead["updated_at"]]
        for ci, val in enumerate(vals, 1):
            if ci == 10: scell(ws1, ri, ci, val)
            elif ci == 11: scell(ws1, ri, ci, val)
            else: dcell(ws1, ri, ci, val, bg=bg)
    for ci, w in enumerate([5,22,14,14,22,14,12,16,14,20,14,30,20,20], 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(h1))}1"

    # ── SHEET 2: CALL LOGS ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Call Logs")
    ws2.sheet_view.showGridLines = False
    ws2.row_dimensions[1].height = 34
    h2 = ["#","Date & Time","Customer","Phone","Product","City","Budget",
          "Telecaller","Call Status","Duration","Followup Date","Remarks"]
    for ci, h in enumerate(h2, 1): hcell(ws2, 1, ci, h, bg="0F172A")
    CBGMAP = {"Connected":"D1FAE5","Not Answered":"FEF9C3","Busy":"FFEDD5",
              "Switched Off":"FEE2E2","Wrong Number":"FEE2E2","Call Back Later":"EDE9FE",
              "Voicemail":"DBEAFE","Interested":"D1FAE5","Not Interested":"FEE2E2",
              "Converted":"D1FAE5","Follow-up":"FFEDD5","Called":"EDE9FE","Not Answered":"FEF9C3"}
    CFGMAP = {"Connected":"065F46","Not Answered":"78350F","Busy":"9A3412",
              "Switched Off":"991B1B","Wrong Number":"991B1B","Call Back Later":"5B21B6",
              "Voicemail":"1E40AF","Interested":"065F46","Not Interested":"991B1B",
              "Converted":"064E3B","Follow-up":"9A3412","Called":"5B21B6"}
    for ri, log in enumerate(call_logs, 2):
        ws2.row_dimensions[ri].height = 20
        bg = "F9FAFB" if ri%2==0 else "FFFFFF"
        vals = [ri-1,log["called_at"],log["lname"],log["lphone"],
                log["lproduct"] or "",log["lcity"] or "",log["lbudget"] or "",
                log["cname"] or "",log["call_status"] or "",
                log["duration"] or "",log["followup_date"] or "",log["remarks"] or ""]
        for ci, val in enumerate(vals, 1): dcell(ws2, ri, ci, val, bg=bg)
        cs = log["call_status"] or ""
        c9 = ws2.cell(row=ri, column=9)
        c9.fill = PatternFill("solid", fgColor=CBGMAP.get(cs, "FFFFFF"))
        c9.font = Font(bold=True, color=CFGMAP.get(cs, "1F2937"), size=10, name="Calibri")
        c9.alignment = Alignment(horizontal="center", vertical="center")
        c9.border = thin
    for ci, w in enumerate([5,20,22,14,24,14,12,18,18,12,14,32], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.freeze_panes = "A2"

    # ── SHEET 3: TELECALLER SUMMARY ─────────────────────────────────────────
    ws3 = wb.create_sheet("Telecaller Summary")
    ws3.sheet_view.showGridLines = False
    ws3.row_dimensions[1].height = 34
    h3 = ["Telecaller","Phone","Total Leads","New","Called","Interested",
          "Not Interested","Follow-up","Converted","Lost","Not Answered",
          "Total Calls","Conversion %"]
    for ci, h in enumerate(h3, 1): hcell(ws3, 1, ci, h, bg="064E3B", fg="D1FAE5")
    db2 = get_db()
    for ri, tc in enumerate(tcs, 2):
        ws3.row_dimensions[ri].height = 22
        uid = tc["id"]
        bg = "F0FDF4" if ri%2==0 else "FFFFFF"
        total = db2.execute("SELECT COUNT(*) FROM leads WHERE assigned_to=?", (uid,)).fetchone()[0]
        calls = db2.execute("SELECT COUNT(*) FROM call_logs WHERE caller_id=?", (uid,)).fetchone()[0]
        stat = {}
        for s in LEAD_STATUSES:
            stat[s] = db2.execute("SELECT COUNT(*) FROM leads WHERE assigned_to=? AND status=?", (uid,s)).fetchone()[0]
        conv = stat.get("Converted",0)
        pct = f"{round(conv/total*100)}%" if total else "0%"
        vals3 = [tc["name"],tc["phone"] or "",total,stat.get("New",0),stat.get("Called",0),
                 stat.get("Interested",0),stat.get("Not Interested",0),stat.get("Follow-up",0),
                 conv,stat.get("Lost",0),stat.get("Not Answered",0),calls,pct]
        for ci, val in enumerate(vals3, 1):
            dcell(ws3, ri, ci, val, bg=bg, align="center" if ci>2 else "left")
        # Highlight interested + converted
        if stat.get("Interested",0) > 0:
            c = ws3.cell(row=ri, column=6)
            c.fill=PatternFill("solid",fgColor="D1FAE5"); c.font=Font(bold=True,color="065F46",size=10,name="Calibri"); c.alignment=Alignment(horizontal="center",vertical="center"); c.border=thin
        if conv > 0:
            c = ws3.cell(row=ri, column=9)
            c.fill=PatternFill("solid",fgColor="D1FAE5"); c.font=Font(bold=True,color="064E3B",size=10,name="Calibri"); c.alignment=Alignment(horizontal="center",vertical="center"); c.border=thin
    db2.close()
    for ci, w in enumerate([22,14,13,10,10,12,16,12,12,10,14,12,13], 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w
    ws3.freeze_panes = "A2"

    out = io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=f"CRM_Report_{today()}.xlsx")

@app.route("/admin/export/csv")
@admin_required
def export_csv():
    db = get_db()
    rows = db.execute("SELECT l.*,u.name as aname FROM leads l LEFT JOIN users u ON l.assigned_to=u.id ORDER BY l.created_at DESC").fetchall()
    db.close()
    out = io.StringIO(); w = csv.writer(out)
    w.writerow(["ID","Name","Phone","WhatsApp","Email","City","Source",
                "Status","Admin Status","Assigned To","Follow-up","Notes","Created"])
    for r in rows:
        w.writerow([r["id"],r["name"],r["phone"],r["whatsapp"] or "",r["email"] or "",
            r["city"] or "",r["source"],r["product"] or "",r["budget"] or "",
            r["status"],r["admin_status"],
            r["aname"] if r["aname"] else "Unassigned",
            r["followup_date"] or "",r["notes"] or "",r["created_at"]])
    out.seek(0)
    return Response(out, mimetype="text/csv",
        headers={"Content-Disposition": f"attachment;filename=leads_{today()}.csv"})

# ══════════════════════════════════════════════════════════════════════════════
#  TELECALLER
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/tc")
@login_required
def tc_dashboard():
    if session["role"] == "admin": return redirect(url_for("admin_dashboard"))
    uid = session["user_id"]; t = today(); db = get_db()
    stats = {
        "total":    db.execute("SELECT COUNT(*) FROM leads WHERE assigned_to=?", (uid,)).fetchone()[0],
        "new":      db.execute("SELECT COUNT(*) FROM leads WHERE assigned_to=? AND status='New'", (uid,)).fetchone()[0],
        "interest": db.execute("SELECT COUNT(*) FROM leads WHERE assigned_to=? AND status='Interested'", (uid,)).fetchone()[0],
        "convert":  db.execute("SELECT COUNT(*) FROM leads WHERE assigned_to=? AND status='Converted'", (uid,)).fetchone()[0],
        "calls":    db.execute("SELECT COUNT(*) FROM call_logs WHERE caller_id=? AND date(called_at)=?", (uid, t)).fetchone()[0],
        "followup": db.execute("SELECT COUNT(*) FROM leads WHERE assigned_to=? AND status='Follow-up'", (uid,)).fetchone()[0],
    }
    leads     = db.execute("SELECT * FROM leads WHERE assigned_to=? AND admin_status='Approved' ORDER BY CASE status WHEN 'Follow-up' THEN 1 WHEN 'New' THEN 2 ELSE 3 END,updated_at DESC LIMIT 15", (uid,)).fetchall()
    followups = db.execute("SELECT * FROM leads WHERE assigned_to=? AND followup_date=?", (uid, t)).fetchall()
    rcalls    = db.execute("SELECT cl.*,l.name as lname,l.phone FROM call_logs cl JOIN leads l ON cl.lead_id=l.id WHERE cl.caller_id=? ORDER BY cl.called_at DESC LIMIT 8", (uid,)).fetchall()
    notifs    = db.execute("SELECT * FROM notifications WHERE user_id=? AND is_read=0 ORDER BY created_at DESC LIMIT 8", (uid,)).fetchall()
    db.close()
    return render_template("tc_dashboard.html", stats=stats, leads=leads,
        followups=followups, rcalls=rcalls, notifs=notifs, today=t, sc=SC)

@app.route("/tc/leads")
@login_required
def tc_leads():
    if session["role"] == "admin": return redirect(url_for("admin_leads"))
    uid = session["user_id"]
    qs = request.args.get("q",""); st = request.args.get("st","")
    page = max(1, int(request.args.get("page",1))); per = 20
    q = "SELECT * FROM leads WHERE assigned_to=? AND admin_status='Approved'"; p = [uid]
    if qs: q += " AND (name LIKE ? OR phone LIKE ? OR city LIKE ?)"; p += [f"%{qs}%"]*3
    if st: q += " AND status=?"; p.append(st)
    db = get_db()
    total  = db.execute(q.replace("SELECT *","SELECT COUNT(*)"), p).fetchone()[0]
    leads  = db.execute(q + f" ORDER BY CASE status WHEN 'Follow-up' THEN 1 WHEN 'New' THEN 2 ELSE 3 END,updated_at DESC LIMIT {per} OFFSET {(page-1)*per}", p).fetchall()
    notifs = db.execute("SELECT * FROM notifications WHERE user_id=? AND is_read=0 ORDER BY created_at DESC LIMIT 8", (uid,)).fetchall()
    db.close()
    return render_template("tc_leads.html", leads=leads, qs=qs, st=st, notifs=notifs,
        statuses=LEAD_STATUSES, sc=SC,
        page=page, total=total, per=per, pages=max(1,(total+per-1)//per), today=today())

@app.route("/tc/leads/<int:lid>")
@login_required
def tc_lead(lid):
    uid = session["user_id"]; db = get_db()
    if session["role"] == "admin":
        lead = db.execute("SELECT l.*,u.name as aname FROM leads l LEFT JOIN users u ON l.assigned_to=u.id WHERE l.id=?", (lid,)).fetchone()
    else:
        lead = db.execute("SELECT * FROM leads WHERE id=? AND assigned_to=?", (lid, uid)).fetchone()
    if not lead: db.close(); flash("Lead nahi mila!", "error"); return redirect(url_for("tc_leads"))
    logs   = db.execute("SELECT cl.*,u.name as cname FROM call_logs cl LEFT JOIN users u ON cl.caller_id=u.id WHERE cl.lead_id=? ORDER BY cl.called_at DESC", (lid,)).fetchall()
    docs   = db.execute("SELECT d.*,u.name as uploader FROM documents d LEFT JOIN users u ON d.uploaded_by=u.id WHERE d.lead_id=? ORDER BY d.uploaded_at DESC", (lid,)).fetchall()
    agents = db.execute("SELECT * FROM users WHERE role='telecaller' AND active=1").fetchall() if session["role"] == "admin" else []
    notifs = db.execute("SELECT * FROM notifications WHERE user_id=? AND is_read=0 ORDER BY created_at DESC LIMIT 8", (uid,)).fetchall()
    db.execute("UPDATE notifications SET is_read=1 WHERE user_id=? AND link LIKE ?", (uid, f"%/leads/{lid}%"))
    db.commit(); db.close()
    # Telecaller gets dedicated template, admin gets full admin template
    if session["role"] == "admin":
        return render_template("lead_detail.html", lead=lead, logs=logs, docs=docs,
            agents=agents, notifs=notifs, statuses=LEAD_STATUSES,
            astatuses=ADMIN_STATUSES, call_statuses=CALL_STATUSES,
            doc_types=DOC_TYPES, doc_statuses=DOC_STATUSES, sc=SC, is_admin=True)
    else:
        return render_template("tc_lead_detail.html", lead=lead, logs=logs, docs=docs,
            notifs=notifs, statuses=LEAD_STATUSES, call_statuses=CALL_STATUSES,
            doc_types=DOC_TYPES, doc_statuses=DOC_STATUSES, sc=SC, is_admin=False,
            now=datetime.now())

@app.route("/tc/leads/<int:lid>/call", methods=["POST"])
@login_required
def log_call(lid):
    uid = session["user_id"]; d = request.form; db = get_db()
    db.execute("INSERT INTO call_logs(lead_id,caller_id,call_status,duration,remarks,followup_date) VALUES(?,?,?,?,?,?)",
        (lid, uid, d.get("call_status",""), d.get("duration",""),
         d.get("remarks",""), d.get("followup_date","") or None))
    ns = d.get("lead_status",""); fu = d.get("followup_date","")
    if ns:
        sq = "UPDATE leads SET status=?,updated_at=datetime('now','localtime')"; sp = [ns]
        if fu: sq += ",followup_date=?"; sp.append(fu)
        db.execute(sq + " WHERE id=?", sp + [lid])
    db.commit(); db.close()
    flash("✅ Call logged!", "success")
    # Redirect: if came from tc/leads list, go back there; otherwise go to detail
    ref = request.referrer or ""
    if session["role"] == "admin":
        return redirect(url_for("admin_lead", lid=lid))
    elif "/tc/leads" in ref and f"/{lid}" not in ref:
        return redirect(url_for("tc_leads"))
    else:
        return redirect(url_for("tc_lead", lid=lid))

# ══════════════════════════════════════════════════════════════════════════════
#  API
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/api/lead/<int:lid>/status", methods=["POST"])
@login_required
def api_status(lid):
    data = request.json
    st = data.get("status")
    if st not in LEAD_STATUSES: return jsonify({"error":"Invalid"}), 400
    followup_date = data.get("followup_date", "")
    note = data.get("note", "")
    db = get_db()
    if followup_date:
        db.execute("UPDATE leads SET status=?,followup_date=?,updated_at=datetime('now','localtime') WHERE id=?", (st, followup_date, lid))
    else:
        db.execute("UPDATE leads SET status=?,updated_at=datetime('now','localtime') WHERE id=?", (st, lid))
    # Auto log in call_logs when status changes via API
    uid = session.get("user_id")
    if uid and st in ["Interested","Not Interested","Converted","Follow-up","Not Answered","Called"]:
        remarks = f"Status changed to {st}" + (f" | {note}" if note else "")
        db.execute("INSERT INTO call_logs(lead_id,caller_id,call_status,remarks,followup_date) VALUES(?,?,?,?,?)",
            (lid, uid, st, remarks, followup_date or None))
    db.commit(); db.close()
    return jsonify({"ok": True, "color": SC.get(st, "#888")})

@app.route("/api/notif/read", methods=["POST"])
@login_required
def notif_read():
    db = get_db()
    db.execute("UPDATE notifications SET is_read=1 WHERE user_id=?", (session["user_id"],))
    db.commit(); db.close()
    return jsonify({"ok": True})

@app.route("/api/notif/count")
@login_required
def notif_count():
    db = get_db()
    n = db.execute("SELECT COUNT(*) FROM notifications WHERE user_id=? AND is_read=0", (session["user_id"],)).fetchone()[0]
    db.close()
    return jsonify({"count": n})

if __name__ == "__main__":
    print("\n" + "="*55)
    print("  🚀 TELECALLER CRM — http://localhost:5000")
    print("  📋 Lead Form   — http://localhost:5000/form")
    print("  🔐 Admin Login — admin / admin123")
    print("="*55 + "\n")
    app.run(debug=True, host="0.0.0.0", port=5000)
